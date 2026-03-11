"""
Gerador dos 5 arquivos Excel de teste para teste_202603111655.ps1

Cenários:
  ex1_tipo_x_simples.xlsx   — 3 lanç. Tipo X,     cols: A=data B=conta_deb C=conta_cred D=valor E=historico
  ex2_tipo_d_rateio.xlsx    — 3 lanç. Tipo D,     cols: A=historico B=valor C=data D=conta_deb E=conta_cred F=tipo G=grupo
  ex3_tipo_c_coleta.xlsx    — 3 lanç. Tipo C,     cols: A=valor B=conta_cred C=conta_deb D=data E=historico F=tipo G=grupo
  ex4_tipo_v_bilateral.xlsx — 4 lanç. Tipo V,     cols: A=data B=tipo C=grupo D=valor E=conta_deb F=conta_cred G=historico
  ex5_misto_xd.xlsx         — 4 lanç. Misto X+D, cols: A=data B=conta_deb C=conta_cred D=valor E=historico F=tipo G=grupo

Contas usadas (mapeamentos esperados):
  3145 → 12   (conta débito padrão)
  2004 → 78   (conta crédito padrão)
  1130 → 25   (conta crédito alternativa)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SAIDA = os.path.dirname(os.path.abspath(__file__))


def _wb_com_cabecalho(headers: list) -> tuple:
    wb = Workbook()
    ws = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    return wb, ws


# ──────────────────────────────────────────────────────────────────
# Excel 1 — Tipo X simples (3 lançamentos)
# Layout: A=data  B=conta_debito  C=conta_credito  D=valor  E=historico
# ──────────────────────────────────────────────────────────────────
def gerar_ex1():
    headers = ["data", "conta_debito", "conta_credito", "valor", "historico"]
    wb, ws = _wb_com_cabecalho(headers)
    rows = [
        ["01/01/2026", "3145", "2004", 60000.00, "VLR ALUGUEL JAN 2026 H.B PARTICIPACOES LTDA"],
        ["02/01/2026", "3145", "1130", 14073.93, "PGTO REF NF. 24536486/1. ENERGISA MATO GROSSO"],
        ["03/01/2026", "3145", "2004",  1438.71, "PAGAMENTO FORNECEDOR ABC SERVICOS"],
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    path = os.path.join(SAIDA, "ex1_tipo_x_simples.xlsx")
    wb.save(path)
    print(f"[OK] {path}")


# ──────────────────────────────────────────────────────────────────
# Excel 2 — Tipo D rateio (1 débito + 2 créditos)
# Layout: A=historico  B=valor  C=data  D=conta_debito  E=conta_credito  F=tipo  G=grupo
# ──────────────────────────────────────────────────────────────────
def gerar_ex2():
    headers = ["historico", "valor", "data", "conta_debito", "conta_credito", "tipo", "grupo"]
    wb, ws = _wb_com_cabecalho(headers)
    rows = [
        # historico                    valor    data          deb    cred   tipo  grupo
        ["RATEIO DESPESA ADM JAN",     500.00, "03/01/2026", "3145", "",    "D", "GRP-D01"],
        ["RATEIO TELECOM PARTE 1",     200.00, "03/01/2026", "",    "2004", "D", "GRP-D01"],
        ["RATEIO TELECOM PARTE 2",     300.00, "03/01/2026", "",    "2004", "D", "GRP-D01"],
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    path = os.path.join(SAIDA, "ex2_tipo_d_rateio.xlsx")
    wb.save(path)
    print(f"[OK] {path}")


# ──────────────────────────────────────────────────────────────────
# Excel 3 — Tipo C coleta (2 débitos + 1 crédito)
# Layout: A=valor  B=conta_credito  C=conta_debito  D=data  E=historico  F=tipo  G=grupo
# ──────────────────────────────────────────────────────────────────
def gerar_ex3():
    headers = ["valor", "conta_credito", "conta_debito", "data", "historico", "tipo", "grupo"]
    wb, ws = _wb_com_cabecalho(headers)
    rows = [
        # valor   cred   deb    data          historico                tipo  grupo
        [200.00, "",    "3145", "06/01/2026", "RECEBIMENTO CLIENTE A", "C", "GRP-C01"],
        [300.00, "",    "3145", "06/01/2026", "RECEBIMENTO CLIENTE B", "C", "GRP-C01"],
        [500.00, "2004", "",   "06/01/2026", "RECEITA CONSOLIDADA",   "C", "GRP-C01"],
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    path = os.path.join(SAIDA, "ex3_tipo_c_coleta.xlsx")
    wb.save(path)
    print(f"[OK] {path}")


# ──────────────────────────────────────────────────────────────────
# Excel 4 — Tipo V bilateral (2 débitos + 2 créditos balanceados)
# Layout: A=data  B=tipo  C=grupo  D=valor  E=conta_debito  F=conta_credito  G=historico
# Σdébitos = 201 = Σcréditos ✓
# ──────────────────────────────────────────────────────────────────
def gerar_ex4():
    headers = ["data", "tipo", "grupo", "valor", "conta_debito", "conta_credito", "historico"]
    wb, ws = _wb_com_cabecalho(headers)
    rows = [
        # data          tipo  grupo      valor   deb    cred   historico
        ["02/01/2026", "V", "GRP-V01", 101.00, "3145", "",    "DEBITO 1 BILATERAL"],
        ["02/01/2026", "V", "GRP-V01", 101.00, "",    "2004", "CREDITO 1 BILATERAL"],
        ["02/01/2026", "V", "GRP-V01", 100.00, "3145", "",    "DEBITO 2 BILATERAL"],
        ["02/01/2026", "V", "GRP-V01", 100.00, "",    "2004", "CREDITO 2 BILATERAL"],
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    path = os.path.join(SAIDA, "ex4_tipo_v_bilateral.xlsx")
    wb.save(path)
    print(f"[OK] {path}")


# ──────────────────────────────────────────────────────────────────
# Excel 5 — Misto X + D (1 X + 3 D) — 4 lançamentos
# Layout: A=data  B=conta_debito  C=conta_credito  D=valor  E=historico  F=tipo  G=grupo
# ──────────────────────────────────────────────────────────────────
def gerar_ex5():
    headers = ["data", "conta_debito", "conta_credito", "valor", "historico", "tipo", "grupo"]
    wb, ws = _wb_com_cabecalho(headers)
    rows = [
        # data          deb    cred   valor    historico                        tipo  grupo
        ["01/01/2026", "3145", "2004", 60000.00, "ALUGUEL JAN 2026",            "X", ""],
        ["02/01/2026", "3145", "",      500.00,  "RATEIO TELECOM DEBITO",       "D", "GRP-M01"],
        ["02/01/2026", "",    "2004",   200.00,  "RATEIO TELECOM CREDITO 1",    "D", "GRP-M01"],
        ["02/01/2026", "",    "2004",   300.00,  "RATEIO TELECOM CREDITO 2",    "D", "GRP-M01"],
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    path = os.path.join(SAIDA, "ex5_misto_xd.xlsx")
    wb.save(path)
    print(f"[OK] {path}")


if __name__ == "__main__":
    gerar_ex1()
    gerar_ex2()
    gerar_ex3()
    gerar_ex4()
    gerar_ex5()
    print("\nTodos os 5 arquivos Excel gerados com sucesso!")
